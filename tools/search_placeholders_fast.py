import os
import glob
import re
import pdfplumber
import openpyxl

PATTERNS = [
    r'\bunknown\b',
    r'\btbd\b',
    r'\btba\b',
    r'\bn/a\b',
    r'\bto be announced\b',
    r'\bto be decided\b',
    r'\bnot assigned\b',
    r'\bunassigned\b',
    r'\bpending\b',
    r'\bplaceholder\b',
    r'\bdummy\b',
    r'\bundefined\b',
    r'\bnull\b',
    r'\bnil\b'
]

REGEX = re.compile('|'.join(PATTERNS), re.IGNORECASE)

def run_scan():
    print("=" * 80)
    print("EXHAUSTIVE PLACEHOLDER SCAN (PDFS & EXCEL DOCUMENTS)")
    print("=" * 80)
    
    findings = []
    
    # 1. PDFs
    pdf_files = sorted(glob.glob("assets/documents/*.pdf"))
    for pf in pdf_files:
        name = os.path.basename(pf)
        with pdfplumber.open(pf) as pdf:
            for p_idx, page in enumerate(pdf.pages):
                text = page.extract_text() or ""
                for line in text.split('\n'):
                    matches = REGEX.findall(line)
                    if matches:
                        findings.append({
                            'file': name,
                            'type': 'PDF Document',
                            'loc': f"Page {p_idx+1}",
                            'matches': list(set(matches)),
                            'text': line.strip()
                        })
                        
    # 2. Excels
    excel_files = [f for f in sorted(glob.glob("assets/documents/*.xlsx")) if not os.path.basename(f).startswith('~$')]
    for ef in excel_files:
        name = os.path.basename(ef)
        wb = openpyxl.load_workbook(ef, data_only=True)
        for sname in wb.sheetnames:
            ws = wb[sname]
            for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                for c_idx, cell in enumerate(row, start=1):
                    if cell is not None:
                        val = str(cell).strip()
                        matches = REGEX.findall(val)
                        if matches:
                            col_letter = openpyxl.utils.get_column_letter(c_idx)
                            findings.append({
                                'file': name,
                                'type': 'Excel Workbook',
                                'loc': f"Sheet '{sname}', Cell {col_letter}{r_idx}",
                                'matches': list(set(matches)),
                                'text': val.replace('\n', ' ')
                            })
                            
    print("\n" + "=" * 80)
    print("AUDIT RESULTS:")
    print("=" * 80)
    
    if not findings:
        print(">> ZERO placeholder terms (unknown, tbd, tba, n/a, dummy, null, etc.) found.")
        print(f">> Scanned {len(pdf_files)} PDF documents and {len(excel_files)} Excel master workbooks.")
    else:
        print(f">> Found {len(findings)} occurrence(s):\n")
        for f in findings:
            print(f"[{f['type']}] {f['file']} ({f['loc']})")
            print(f"  Matched: {f['matches']}")
            print(f"  Excerpt: \"{f['text']}\"\n")

if __name__ == '__main__':
    run_scan()
