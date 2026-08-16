import os
import glob
import pdfplumber
import openpyxl

artifact_dir = r"C:\Users\MALIK\.gemini\antigravity-ide\brain\83f6734c-1364-4521-80ff-f595797d00c4"

def render_real_pdf_pages():
    print("==================================================")
    print("RENDERING NATIVE PDF PAGES DIRECTLY TO SCREENSHOTS")
    print("==================================================")
    
    pdf_files = [
        "CE-1.pdf",
        "CS-12-1.pdf",
        "EE.pdf",
        "FSn,BTY,BCH,HND,RBS.pdf",
        "HUM-1.pdf",
        "ME-1.pdf",
        "MS-2.pdf"
    ]
    
    for pf_name in pdf_files:
        pf_path = os.path.join("assets/documents", pf_name)
        if not os.path.exists(pf_path):
            continue
            
        with pdfplumber.open(pf_path) as pdf:
            # Render page 1 of each PDF
            p1 = pdf.pages[0]
            out_img_name = f"native_pdf_{pf_name.replace('.pdf', '').replace(',', '_')}_p1.png"
            out_path = os.path.join(artifact_dir, out_img_name)
            img = p1.to_image(resolution=200)
            img.save(out_path)
            print(f"Saved Native PDF Page: {out_img_name}")

def render_authentic_html_excel(excel_path, out_name):
    print(f"\nBuilding Authentic Excel Sheet HTML Viewer for {os.path.basename(excel_path)}...")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    
    html = """<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<style>
  body {
    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
    background-color: #f3f3f3;
    margin: 0;
    padding: 20px;
  }
  .excel-window {
    background: #fff;
    border: 1px solid #d4d4d4;
    box-shadow: 0 4px 12px rgba(0,0,0,0.15);
    border-radius: 4px;
    overflow-x: auto;
    max-width: 100%;
  }
  .excel-ribbon {
    background: #107c41;
    color: #fff;
    padding: 10px 16px;
    font-size: 15px;
    font-weight: 600;
    display: flex;
    align-items: center;
    gap: 12px;
  }
  .excel-formula-bar {
    background: #f9f9f9;
    border-bottom: 1px solid #e1e1e1;
    padding: 6px 12px;
    font-size: 12px;
    color: #444;
    display: flex;
    gap: 10px;
  }
  table.excel-grid {
    border-collapse: collapse;
    width: 100%;
    font-size: 11px;
    table-layout: auto;
  }
  table.excel-grid th, table.excel-grid td {
    border: 1px solid #d4d4d4;
    padding: 4px 6px;
    white-space: nowrap;
    text-overflow: ellipsis;
  }
  table.excel-grid th {
    background: #e6e6e6;
    color: #333;
    font-weight: 600;
    text-align: center;
    user-select: none;
  }
  .row-num {
    background: #e6e6e6;
    color: #555;
    font-weight: 600;
    text-align: center;
    width: 35px;
    min-width: 35px;
  }
  .title-row {
    background: #f2f8f4;
    font-weight: 700;
    font-size: 13px;
    color: #107c41;
    text-align: center;
  }
  .header-room {
    background: #e8f0fe;
    color: #1a73e8;
    font-weight: 700;
    text-align: center;
  }
  .batch-cell {
    background: #fff;
    color: #1e293b;
    font-weight: 700;
  }
  .compound-batch {
    background: #fef3c7;
    color: #92400e;
    font-weight: 700;
  }
  .subject-cell {
    background: #fafafa;
    color: #047857;
  }
  .date-cell {
    background: #f0fdf4;
    font-weight: 700;
    color: #166534;
  }
  .time-cell {
    background: #f8fafc;
    font-weight: 600;
    color: #0284c7;
  }
</style>
</head>
<body>
<div class="excel-window">
  <div class="excel-ribbon">
    <span>📊 Excel 365 // """ + os.path.basename(excel_path) + """</span>
  </div>
  <div class="excel-formula-bar">
    <span style="font-weight: 700; color: #107c41;">fx</span>
    <span>=COMSATS_UNIVERSITY_OFFICIAL_DATE_SHEET_GRID()</span>
  </div>
  <table class="excel-grid">
    <thead>
      <tr>
        <th class="row-num"></th>
"""
    # Column headers A, B, C...
    num_cols = len(rows[2]) if len(rows) > 2 else 53
    for c in range(num_cols):
        col_letter = openpyxl.utils.get_column_letter(c + 1)
        html += f"        <th>{col_letter}</th>\n"
    html += "      </tr>\n    </thead>\n    <tbody>\n"
    
    # Rows
    for r_idx, row in enumerate(rows[:35]):
        row_num = r_idx + 1
        html += f"      <tr>\n        <td class=\"row-num\">{row_num}</td>\n"
        for c_idx in range(num_cols):
            val = row[c_idx] if c_idx < len(row) else None
            s_val = str(val).strip() if val is not None else ""
            
            # CSS styling
            cell_class = ""
            if row_num == 2:
                cell_class = "title-row"
            elif row_num == 3:
                cell_class = "header-room"
            elif row_num >= 4:
                if c_idx in [0, 12, 26, 41]:
                    cell_class = "date-cell"
                elif c_idx in [1, 13, 27, 42]:
                    cell_class = "time-cell"
                else:
                    if row_num % 3 == 1 or (row_num % 3 == 2 and 'FA' in s_val or 'SP' in s_val):
                        if '-' in s_val and s_val.count('-') >= 3:
                            cell_class = "compound-batch"
                        else:
                            cell_class = "batch-cell"
                    else:
                        cell_class = "subject-cell"
                        
            html += f"        <td class=\"{cell_class}\">{s_val}</td>\n"
        html += "      </tr>\n"
        
    html += """    </tbody>
  </table>
</div>
</body>
</html>"""

    html_path = os.path.join(artifact_dir, f"{out_name}.html")
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"Saved HTML Excel view: {html_path}")
    return html_path

if __name__ == '__main__':
    render_real_pdf_pages()
    f1 = "assets/documents/Date_Sheet_FINAL_Term_Exam_FALL_2025 - Version I - 08-12-25.xlsx"
    f2 = "assets/documents/Version I of Date Sheet Mid Term Exam SPRING 2026 25-03-2026.xlsx"
    render_authentic_html_excel(f1, "excel_authentic_final_fall_2025")
    render_authentic_html_excel(f2, "excel_authentic_mid_spring_2026")
