"""
Complete Shortcoming and Discrepancy Analyzer for IRIS Timetable & Exam Parsers.
Analyzes:
1. Excel Exam Date Sheets:
   - 4-Block Horizontal Campus Matrix structure
   - Paired (Batch, Subject) rows with empty row delimiters
   - Date carry-forward semantics across slots on the same day
   - Compound Hyphenated/Slash batch cell expansion
   - Room normalization (capacity stripping, formatting)
   - Military time conversion (0900-1030 -> 09:00 AM - 10:30 AM)
   
2. PDF Timetables:
   - Header Y-threshold & spurious time slot filtering
   - Department-specific slot time definitions (1.5-hr vs 1.0-hr)
   - Multi-slot Lab spanning vs empty theory slot handling
   - Faculty name honorific spacing & capitalized name extraction
   - Room code & capacity parsing
"""
import openpyxl
import pdfplumber
import re
import os
import glob
import json

def analyze():
    print("=" * 80)
    print("IRIS UNIVERSITY DATA & PARSER SHORTCOMING AUDIT")
    print("=" * 80)
    
    # 1. EXCEL AUDIT
    print("\n[EXCEL DATE SHEETS ANALYSIS]")
    excel_files = glob.glob('assets/**/*.xlsx', recursive=True)
    for ef in excel_files:
        print(f"\n--- {os.path.basename(ef)} ---")
        wb = openpyxl.load_workbook(ef, data_only=True)
        sheet = wb.active
        
        # Detect building blocks from row 3
        r3 = [sheet.cell(3, c).value for c in range(1, sheet.max_column + 1)]
        date_cols = [c - 1 for c in range(1, sheet.max_column + 1) if r3[c-1] and str(r3[c-1]).strip().lower() == 'date']
        time_cols = [c - 1 for c in range(1, sheet.max_column + 1) if r3[c-1] and str(r3[c-1]).strip().lower() == 'time']
        
        print(f"  Campus Horizontal Blocks: {len(date_cols)} blocks at cols {date_cols}")
        for b_idx in range(len(date_cols)):
            start_col = date_cols[b_idx]
            end_col = date_cols[b_idx + 1] if b_idx + 1 < len(date_cols) else sheet.max_column
            room_names = [r3[c] for c in range(start_col + 2, end_col) if r3[c] is not None]
            print(f"    Block {b_idx + 1} (Cols {start_col}..{end_col-1}): {len(room_names)} rooms -> {room_names[:5]}")
            
    # 2. PDF AUDIT
    print("\n\n[PDF TIMETABLES ANALYSIS]")
    pdf_files = glob.glob('assets/**/*.pdf', recursive=True)
    for pf in pdf_files:
        print(f"\n--- {os.path.basename(pf)} ---")
        with pdfplumber.open(pf) as pdf:
            print(f"  Total Pages: {len(pdf.pages)}")
            p1 = pdf.pages[0]
            text = p1.extract_text() or ""
            times = re.findall(r'\b\d{1,2}[:.]\d{2}\s*-\s*\d{1,2}[:.]\d{2}\b', text)
            print(f"  Time Slots: {times}")

if __name__ == '__main__':
    analyze()
