import os
import glob
import json
import pdfplumber
import openpyxl
import re

print("=" * 70)
print("COMPREHENSIVE AUDIT OF ALL PDFS AND EXCELS IN REPO")
print("=" * 70)

# 1. Inspect all PDFs in assets
pdf_files = glob.glob('assets/**/*.pdf', recursive=True)
print(f"\n[1] FOUND {len(pdf_files)} PDF FILES:")
for p in pdf_files:
    print(f"  - {p} (size: {os.path.getsize(p)} bytes)")

# 2. Inspect all Excels in assets
excel_files = glob.glob('assets/**/*.xlsx', recursive=True) + glob.glob('assets/**/*.xls', recursive=True)
print(f"\n[2] FOUND {len(excel_files)} EXCEL FILES:")
for e in excel_files:
    print(f"  - {e} (size: {os.path.getsize(e)} bytes)")

# 3. Deep dive into each PDF structure
print("\n[3] PDF DETAILED STRUCTURAL & TEXT AUDIT:")
for p in pdf_files:
    print(f"\n--- Analyzing PDF: {p} ---")
    with pdfplumber.open(p) as pdf:
        print(f"  Pages count: {len(pdf.pages)}")
        for idx, page in enumerate(pdf.pages):
            text = page.extract_text() or ""
            words = page.extract_words()
            tables = page.extract_tables()
            print(f"  Page {idx+1}: {len(words)} words, {len(tables)} tables extracted, text length: {len(text)}")
            first_lines = [l.strip() for l in text.split("\n") if l.strip()][:5]
            print(f"    Header preview: {' | '.join(first_lines)}")

# 4. Deep dive into each Excel structure
print("\n[4] EXCEL DETAILED SHEET & DATA AUDIT:")
for e in excel_files:
    print(f"\n--- Analyzing Excel: {e} ---")
    try:
        wb = openpyxl.load_workbook(e, data_only=True)
        print(f"  Sheet names: {wb.sheetnames}")
        for sname in wb.sheetnames:
            sheet = wb[sname]
            rows = list(sheet.iter_rows(values_only=True))
            print(f"  Sheet '{sname}': {len(rows)} rows, {len(rows[0]) if rows else 0} cols")
            # find first 5 non-empty rows
            non_empty = [r for r in rows if any(cell is not None for cell in r)][:5]
            print(f"    Preview header:")
            for r in non_empty:
                print(f"      {r[:8]}")
    except Exception as ex:
        print(f"  Error reading {e}: {ex}")
