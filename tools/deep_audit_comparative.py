import os
import glob
import re
import openpyxl
import pdfplumber

def audit_all():
    print("=" * 80)
    print("DEEP COMPARATIVE AUDIT: UNIVERSITY PDFS & EXCEL DATE SHEETS")
    print("=" * 80)

    # ----------------------------------------------------
    # 1. EXCEL EXAM DATE SHEETS AUDIT
    # ----------------------------------------------------
    excel_files = glob.glob('assets/**/*.xlsx', recursive=True) + glob.glob('assets/**/*.xls', recursive=True)
    print("\n>>> AUDITING EXCEL DATE SHEETS:")
    
    for ef in excel_files:
        print(f"\n[FILE]: {os.path.basename(ef)}")
        wb = openpyxl.load_workbook(ef, data_only=True)
        for sname in wb.sheetnames:
            sheet = wb[sname]
            all_rows = list(sheet.iter_rows(values_only=True))
            print(f"  Sheet '{sname}': {len(all_rows)} total rows")
            
            # Find header row
            header_idx = -1
            for idx, r in enumerate(all_rows[:15]):
                r_str = " ".join([str(c) for c in r if c is not None]).lower()
                if 'course' in r_str or 'subject' in r_str or 'date' in r_str or 'class' in r_str or 'room' in r_str:
                    header_idx = idx
                    print(f"    Detected Header Row at Row {idx+1}: {r[:10]}")
                    break
            
            if header_idx != -1:
                col_headers = [str(c).strip() if c is not None else "" for c in all_rows[header_idx]]
                print(f"    Columns ({len(col_headers)}): {col_headers}")
                
                # Check data anomalies
                row_count = 0
                anomalies = []
                multi_class_cells = []
                date_formats = set()
                time_formats = set()
                
                for r_idx in range(header_idx + 1, len(all_rows)):
                    row = all_rows[r_idx]
                    if not any(row): continue
                    row_count += 1
                    
                    row_dict = {col_headers[i]: row[i] for i in range(min(len(col_headers), len(row)))}
                    
                    # Check class / batch cell
                    class_val = str(row_dict.get('Class', row_dict.get('Batch', row_dict.get('Program', '')))).strip()
                    if '/' in class_val or '&' in class_val or ',' in class_val or '\n' in class_val:
                        multi_class_cells.append((r_idx+1, class_val))
                    
                    # Check date and time
                    date_val = str(row_dict.get('Date', '')).strip()
                    time_val = str(row_dict.get('Time', row_dict.get('Slot', ''))).strip()
                    if date_val: date_formats.add(date_val[:10])
                    if time_val: time_formats.add(time_val)
                    
                    # Check missing course or room
                    course_val = str(row_dict.get('Course Title', row_dict.get('Subject', row_dict.get('Course', '')))).strip()
                    room_val = str(row_dict.get('Room #', row_dict.get('Room', row_dict.get('Room No', '')))).strip()
                    
                    if not course_val or course_val == 'None' or course_val == '':
                        anomalies.append(f"Row {r_idx+1}: Missing Course Title ({row})")
                    if not room_val or room_val == 'None' or room_val == '':
                        anomalies.append(f"Row {r_idx+1}: Missing Room ({row})")
                
                print(f"    Valid Exam Rows: {row_count}")
                print(f"    Multi-Batch Merged Cells Count: {len(multi_class_cells)}")
                if multi_class_cells:
                    print(f"      Samples: {multi_class_cells[:4]}")
                print(f"    Distinct Time Formats: {list(time_formats)[:5]}")
                print(f"    Anomalies Count: {len(anomalies)}")
                if anomalies:
                    print(f"      Sample Anomalies: {anomalies[:3]}")

    # ----------------------------------------------------
    # 2. PDF TIMETABLES AUDIT
    # ----------------------------------------------------
    pdf_files = glob.glob('assets/**/*.pdf', recursive=True)
    print("\n\n>>> AUDITING PDF TIMETABLES:")
    
    for pf in pdf_files:
        fname = os.path.basename(pf)
        print(f"\n[FILE]: {fname}")
        with pdfplumber.open(pf) as pdf:
            print(f"  Pages: {len(pdf.pages)}")
            
            total_cells = 0
            empty_cells = 0
            two_line_cells = 0
            three_line_cells = 0
            four_plus_cells = 0
            room_patterns = set()
            teacher_samples = set()
            batch_samples = set()
            unparsed_potential_issues = []
            
            for p_idx, page in enumerate(pdf.pages):
                tables = page.extract_tables()
                if not tables:
                    # Let's inspect raw text tokens
                    words = page.extract_words()
                    print(f"  Page {p_idx+1}: No standard table lines, extracted {len(words)} raw text tokens")
                    continue
                
                for t in tables:
                    # Row 0 usually headers
                    for r_idx, row in enumerate(t):
                        if r_idx == 0: continue
                        for c_idx, cell in enumerate(row):
                            if c_idx == 0:
                                # Batch column
                                if cell:
                                    batch_samples.add(cell.replace("\n", " ").strip())
                                continue
                            
                            if not cell or cell.strip() == '':
                                empty_cells += 1
                                continue
                            
                            total_cells += 1
                            lines = [l.strip() for l in cell.split("\n") if l.strip()]
                            if len(lines) == 1:
                                unparsed_potential_issues.append((p_idx+1, r_idx, c_idx, f"Single line cell: '{lines[0]}'"))
                            elif len(lines) == 2:
                                two_line_cells += 1
                            elif len(lines) == 3:
                                three_line_cells += 1
                            else:
                                four_plus_cells += 1
                                
                            # Check room tokens
                            last_line = lines[-1] if lines else ""
                            if re.search(r'\b(?:Lab|Room|CR|Hall|LH|[A-D]\d|C\d\.\d)\b', last_line, re.I):
                                room_patterns.add(last_line)
                            
                            # Check teacher tokens
                            if len(lines) >= 2:
                                teacher_samples.add(lines[0])

            print(f"  Non-Empty Class Cells: {total_cells} (2-lines: {two_line_cells}, 3-lines: {three_line_cells}, 4+ lines: {four_plus_cells})")
            print(f"  Sample Batches ({len(batch_samples)}): {list(batch_samples)[:4]}")
            print(f"  Potential Cell Anomalies ({len(unparsed_potential_issues)}):")
            for issue in unparsed_potential_issues[:5]:
                print(f"    Page {issue[0]} Row {issue[1]} Col {issue[2]}: {issue[3]}")

if __name__ == '__main__':
    audit_all()
