import openpyxl
import os

def dump_raw_excel(path):
    print("=" * 80)
    print(f"RAW EXCEL DUMP: {os.path.basename(path)}")
    print("=" * 80)
    wb = openpyxl.load_workbook(path, data_only=False)
    for sname in wb.sheetnames:
        sheet = wb[sname]
        print(f"\n--- SHEET: '{sname}' ---")
        print(f"Dimensions: {sheet.dimensions}")
        print(f"Merged Cell Ranges ({len(sheet.merged_cells.ranges)}):")
        for m in list(sheet.merged_cells.ranges)[:15]:
            print(f"  Merged: {m}")
            
        print("\nFirst 12 Raw Rows (all columns):")
        for r_idx in range(1, 13):
            row_vals = [sheet.cell(r_idx, c_idx).value for c_idx in range(1, sheet.max_column + 1)]
            non_empty_indices = [i for i, v in enumerate(row_vals) if v is not None]
            if non_empty_indices:
                print(f"Row {r_idx:2d} ({len(non_empty_indices)} cells):")
                for i in non_empty_indices:
                    col_letter = openpyxl.utils.get_column_letter(i + 1)
                    print(f"    [{col_letter}{r_idx} = col {i:2d}]: {repr(row_vals[i])}")
            else:
                print(f"Row {r_idx:2d}: EMPTY")

dump_raw_excel("assets/documents/Date_Sheet_FINAL_Term_Exam_FALL_2025 - Version I - 08-12-25.xlsx")
dump_raw_excel("assets/documents/Version I of Date Sheet Mid Term Exam SPRING 2026 25-03-2026.xlsx")
