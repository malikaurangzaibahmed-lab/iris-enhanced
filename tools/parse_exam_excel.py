import os
import sys
import json
import openpyxl
import re

def parse_exam_excel(path):
    print(f"=== Parsing Date Sheet: {os.path.basename(path)} ===")
    if not os.path.exists(path):
        print(f"Error: File not found {path}")
        return []
        
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        sheet = wb.active
        print(f"Sheet: {sheet.title} ({sheet.max_row} rows, {sheet.max_column} cols)")
        
        # Row 3 (1-indexed index 3) contains headers
        # Col 1: Date, Col 2: Time, Col 3+: Rooms
        header_row = 3
        rooms = []
        for col in range(3, sheet.max_column + 1):
            val = sheet.cell(header_row, col).value
            if val:
                rooms.append((col, str(val).strip()))
                
        print(f"Detected {len(rooms)} exam rooms/venues.")
        
        exams = []
        current_date = None
        current_time = None
        
        # Even/Odd rows pairings after header:
        # Row R: Batch info
        # Row R+1: Subject info
        r = 4
        while r <= sheet.max_row:
            row_date = sheet.cell(r, 1).value
            row_time = sheet.cell(r, 2).value
            
            row_date_str = str(row_date).strip() if row_date is not None else ""
            row_time_str = str(row_time).strip() if row_time is not None else ""
            
            # Check if it's a repeated header
            if row_date_str.lower() == "date" or row_time_str.lower() == "time":
                r += 1
                continue
                
            # Check if it's an empty row
            if not row_time_str:
                r += 1
                continue
                
            # Valid exam row
            if row_date_str:
                current_date = row_date_str
            current_time = row_time_str
            
            # Scan each room column
            for col_idx, room_name in rooms:
                batch_cell = sheet.cell(r, col_idx).value
                subject_cell = sheet.cell(r + 1, col_idx).value
                
                if batch_cell is not None and subject_cell is not None:
                    batch_str = str(batch_cell).strip()
                    subject_str = str(subject_cell).strip()
                    
                    if not batch_str or not subject_str:
                        continue
                        
                    # Ignore header repetitions
                    if batch_str.lower() in ("date", "time") or subject_str.lower() in ("date", "time"):
                        continue
                    if batch_str == room_name:
                        continue
                    if any(room[1] == batch_str for room in rooms):
                        continue
                    
                    batches = split_combined_cell(batch_str)
                    subjects = split_combined_cell(subject_str)
                    
                    max_len = max(len(batches), len(subjects))
                    for idx in range(max_len):
                        b = batches[idx] if idx < len(batches) else batches[-1]
                        s = subjects[idx] if idx < len(subjects) else subjects[-1]
                        
                        exams.append({
                            "date": current_date or "Unknown Date",
                            "time": current_time,
                            "room": room_name,
                            "batch": b,
                            "subject": s
                        })
                        
            r += 2  # Move to next pair (Batch + Subject)
            
        print(f"Successfully extracted {len(exams)} individual exam slots.")
        
        # Save output json
        out_filename = os.path.basename(path).replace(".xlsx", "_parsed.json")
        out_path = os.path.join(os.path.dirname(path), out_filename)
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(exams, f, indent=2)
            
        print(f"Saved parsed JSON to: {out_path}")
        return exams
    except Exception as e:
        print(f"Failed parsing exam sheet: {e}")
        import traceback
        traceback.print_exc()
        return []

def split_combined_cell(val):
    if not val:
        return []
    
    val = val.strip().rstrip('-')
    parts = re.split(r'-(?=FA\d{2}|SP\d{2})', val, flags=re.IGNORECASE)
    
    result = []
    for part in parts:
        part = part.strip()
        m = re.match(r'^((?:FA|SP)\d{2}-[A-Z0-9]+)(?:-([A-Z0-9,\s/&]+))?$', part, re.IGNORECASE)
        if m:
            base = m.group(1)
            suffix = m.group(2)
            if suffix:
                sections = re.split(r'[,/&]', suffix)
                sections = [s.strip() for s in sections if s.strip()]
                if all(len(s) <= 3 for s in sections):
                    for s in sections:
                        result.append(f"{base}-{s}")
                else:
                    result.append(part)
            else:
                result.append(base)
        else:
            alt_parts = re.split(r'[,/]', part)
            result.extend([p.strip() for p in alt_parts if p.strip()])
            
    return [r.strip().rstrip('-') for r in result if r.strip()]

if __name__ == "__main__":
    assets_dir = r"d:\Ai models\IRIS\assets"
    f1 = os.path.join(assets_dir, "Date_Sheet_FINAL_Term_Exam_FALL_2025 - Version I - 08-12-25.xlsx")
    f2 = os.path.join(assets_dir, "Version I of Date Sheet Mid Term Exam SPRING 2026 25-03-2026.xlsx")
    
    if os.path.exists(f1):
        parse_exam_excel(f1)
    if os.path.exists(f2):
        parse_exam_excel(f2)
