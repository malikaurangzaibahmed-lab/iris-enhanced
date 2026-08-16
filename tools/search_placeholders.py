import os
import glob
import re
import pdfplumber
import openpyxl

PLACEHOLDER_PATTERNS = [
    r'\bunknown\b',
    r'\btbd\b',
    r'\btba\b',
    r'\bn/a\b',
    r'\bto be announced\b',
    r'\bto be decided\b',
    r'\bnot assigned\b',
    r'\bunassigned\b',
    r'\bpending\b',
    r'\bplaceholder\b',
    r'\bdummy\b',
    r'\bundefined\b',
    r'\bnull\b',
    r'\bnil\b',
    r'\btemp\b',
    r'\btest\b'
]

REGEX = re.compile('|'.join(PLACEHOLDER_PATTERNS), re.IGNORECASE)

def scan_documents():
    print("=" * 80)
    print("SEARCHING FOR PLACEHOLDER TERMS (UNKNOWN, TBD, TBA, N/A, ETC.)")
    print("=" * 80)
    
    findings = []
    
    # 1. Scan PDFs
    pdf_files = sorted(glob.glob("assets/documents/*.pdf"))
    for pf in pdf_files:
        name = os.path.basename(pf)
        print(f"\nScanning PDF: {name}...")
        with pdfplumber.open(pf) as pdf:
            for p_idx, page in enumerate(pdf.pages):
                text = page.extract_text() or ""
                tables = page.extract_tables()
                
                # Check raw page text
                for line_idx, line in enumerate(text.split('\n')):
                    matches = REGEX.findall(line)
                    if matches:
                        # Filter out legitimate substrings e.g. "Contemp", "Contemporary", "Latest", "Testing" if part of legitimate course
                        valid_matches = []
                        for m in matches:
                            # Avoid false positive on "Contemporary" or "Software Testing"
                            lower_line = line.lower()
                            if m.lower() == 'test' and ('software testing' in lower_line or 'testing' in lower_line):
                                continue
                            if m.lower() == 'temp' and ('contemporary' in lower_line or 'attempt' in lower_line or 'temperature' in lower_line):
                                continue
                            valid_matches.append(m)
                        if valid_matches:
                            findings.append({
                                'file': name,
                                'type': 'PDF',
                                'location': f"Page {p_idx + 1}, Line {line_idx + 1}",
                                'matches': valid_matches,
                                'context': line.strip()
                            })
                            
                # Check table cells
                for t_idx, table in enumerate(tables):
                    for r_idx, row in enumerate(table):
                        for c_idx, cell in enumerate(row):
                            if cell:
                                cell_str = str(cell).strip()
                                matches = REGEX.findall(cell_str)
                                valid_matches = []
                                for m in matches:
                                    lower_cell = cell_str.lower()
                                    if m.lower() == 'test' and 'software testing' in lower_cell:
                                        continue
                                    if m.lower() == 'temp' and 'contemporary' in lower_cell:
                                        continue
                                    valid_matches.append(m)
                                if valid_matches:
                                    findings.append({
                                        'file': name,
                                        'type': 'PDF Table Cell',
                                        'location': f"Page {p_idx + 1}, Table {t_idx + 1}, Row {r_idx + 1}, Col {c_idx + 1}",
                                        'matches': valid_matches,
                                        'context': cell_str.replace('\n', ' ')
                                    })
                                    
    # 2. Scan Excel files
    excel_files = sorted(glob.glob("assets/documents/*.xlsx"))
    for ef in excel_files:
        name = os.path.basename(ef)
        print(f"\nScanning Excel: {name}...")
        wb = openpyxl.load_workbook(ef, data_only=True)
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                for col_idx, cell_val in enumerate(row, start=1):
                    if cell_val is not None:
                        val_str = str(cell_val).strip()
                        matches = REGEX.findall(val_str)
                        valid_matches = []
                        for m in matches:
                            lower_val = val_str.lower()
                            if m.lower() == 'test' and 'software testing' in lower_val:
                                continue
                            if m.lower() == 'temp' and 'contemporary' in lower_val:
                                continue
                            valid_matches.append(m)
                        if valid_matches:
                            col_letter = openpyxl.utils.get_column_letter(col_idx)
                            findings.append({
                                'file': name,
                                'type': 'Excel Cell',
                                'location': f"Sheet '{sheetname}', Cell {col_letter}{row_idx}",
                                'matches': valid_matches,
                                'context': val_str.replace('\n', ' ')
                            })
                            
    print("\n" + "=" * 80)
    print("SEARCH REPORT SUMMARY")
    print("=" * 80)
    if not findings:
        print("ZERO placeholder terms found across all 8 PDFs and 2 Excel workbooks!")
    else:
        print(f"Found {len(findings)} potential occurrences:\n")
        for f in findings:
            print(f"[{f['type']}] {f['file']} -> {f['location']}")
            print(f"  Matched Terms: {', '.join(set(f['matches']))}")
            print(f"  Exact Content: \"{f['context']}\"")
            print("-" * 60)

if __name__ == '__main__':
    scan_documents()
