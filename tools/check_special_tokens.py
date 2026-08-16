import os
import glob
import re
import pdfplumber
import openpyxl

def check_special_faculty_and_venues():
    print("=" * 80)
    print("SCANNING FOR VISITING FACULTY, STAFF, AND SPECIAL TOKENS")
    print("=" * 80)
    
    visiting_faculty = set()
    special_venues = set()
    
    # 1. PDFs
    pdf_files = sorted(glob.glob("assets/documents/*.pdf"))
    for pf in pdf_files:
        with pdfplumber.open(pf) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                # Find all VS / Visiting / Staff entries
                vs_matches = re.findall(r'\b(?:VS|Visiting|Staff|TBA|TBD)\s+[A-Za-z\.\s]+', text)
                for m in vs_matches:
                    clean = m.split('\n')[0].strip()
                    if clean: visiting_faculty.add(clean)
                    
    # 2. Excels
    excel_files = [f for f in sorted(glob.glob("assets/documents/*.xlsx")) if not os.path.basename(f).startswith('~$')]
    for ef in excel_files:
        wb = openpyxl.load_workbook(ef, data_only=True)
        for sname in wb.sheetnames:
            ws = wb[sname]
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell:
                        s = str(cell).strip()
                        if any(k in s.lower() for k in ['hall', 'auditorium', 'lab', 'ground', 'workshop']):
                            special_venues.add(s)
                            
    print(f"\nVisiting Faculty (VS) Detected: {len(visiting_faculty)}")
    for vf in sorted(visiting_faculty)[:10]:
        print(f"  - {vf}")
        
    print(f"\nSpecial Examination Venues & Labs in Date Sheets: {len(special_venues)}")
    for sv in sorted(special_venues)[:10]:
        print(f"  - {sv}")

if __name__ == '__main__':
    check_special_faculty_and_venues()
