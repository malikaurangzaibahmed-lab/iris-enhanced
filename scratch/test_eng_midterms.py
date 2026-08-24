import re
from test_all_batches_coverage import is_batch_match, parse_batch_key, expand_batch_sections

# Test engineering batches
eng_batches = ['FA24-BEE-4-A', 'FA25-BME-2-A', 'FA24-CVE-4-A', 'FA23-CVE-6-A', 'FA22-BME-8-A']

# Run extractor on midterms
import openpyxl
file_path = 'assets/documents/Version I of Date Sheet Mid Term Exam SPRING 2026 25-03-2026.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb.active

venue_cols = {}
cur_date_col = -1
cur_time_col = -1
for c in range(1, sheet.max_column + 1):
    val = str(sheet.cell(3, c).value or '').strip()
    if 'DATE' in val.upper(): cur_date_col = c
    elif 'TIME' in val.upper() or 'SLOT' in val.upper(): cur_time_col = c
    elif val:
        venue_cols[c] = {'venue': re.sub(r'\s*\(\d+\)', '', val).strip(), 'date_col': cur_date_col, 'time_col': cur_time_col}

all_exams = []
active_dates = {}
active_times = {}
r = 4
while r <= sheet.max_row:
    row_vals = [str(sheet.cell(r, c).value or '').strip() for c in range(1, sheet.max_column + 1)]
    if not any(row_vals):
        r += 1
        continue
    row_upper = [v.upper() for v in row_vals if v]
    if any('DATE' in v for v in row_upper) and any('TIME' in v for v in row_upper):
        r += 1
        continue
    for d_col in set(info['date_col'] for info in venue_cols.values() if info['date_col'] != -1):
        d_val = str(sheet.cell(r, d_col).value or '').strip()
        if d_val and 'DATE' not in d_val.upper(): active_dates[d_col] = d_val
    for t_col in set(info['time_col'] for info in venue_cols.values() if info['time_col'] != -1):
        t_val = str(sheet.cell(r, t_col).value or '').strip()
        if t_val and 'TIME' not in t_val.upper(): active_times[t_col] = t_val
    subject_row = r + 1
    for c, info in venue_cols.items():
        b_val = str(sheet.cell(r, c).value or '').strip()
        s_val = str(sheet.cell(subject_row, c).value or '').strip()
        if not b_val or not s_val or b_val.upper() in ['BATCH', 'CLASS', 'DATE', 'TIME'] or s_val.upper() in ['SUBJECT', 'COURSE', 'DATE', 'TIME']:
            continue
        all_exams.append({
            'date': active_dates.get(info['date_col'], 'TBD'),
            'time': active_times.get(info['time_col'], '09:00 AM - 10:30 AM'),
            'room': info['venue'],
            'batch': b_val,
            'subject': s_val
        })
    r += 2

for eb in eng_batches:
    matched = [e for e in all_exams if is_batch_match(eb, e['batch'])]
    unique_papers = sorted(list(set(f"{e['date']} - {e['subject']}" for e in matched)))
    print(f"\n{eb} -> {len(matched)} entries, {len(unique_papers)} unique papers:")
    for p in unique_papers:
        print(f"    * {p}")
