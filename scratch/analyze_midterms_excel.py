import openpyxl
import re

file_path = 'assets/documents/Version I of Date Sheet Mid Term Exam SPRING 2026 25-03-2026.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb.active

print(f"Sheet Name: {sheet.title}, Max Row: {sheet.max_row}, Max Col: {sheet.max_column}")

# Print first 6 rows
for r in range(1, 7):
    row_vals = [str(sheet.cell(r, c).value or '').strip() for c in range(1, min(sheet.max_column+1, 20))]
    print(f"R{r}: {row_vals}")

# Identify venue columns in row 3
venue_cols = {}
cur_date_col = -1
cur_time_col = -1

for c in range(1, sheet.max_column + 1):
    val = str(sheet.cell(3, c).value or '').strip()
    val_upper = val.upper()
    if 'DATE' in val_upper:
        cur_date_col = c
    elif 'TIME' in val_upper or 'SLOT' in val_upper:
        cur_time_col = c
    elif val:
        venue_clean = re.sub(r'\s*\(\d+\)', '', val).strip()
        venue_cols[c] = {
            'venue': venue_clean,
            'date_col': cur_date_col,
            'time_col': cur_time_col
        }

print(f"\nVenue columns mapped: {len(venue_cols)}")

all_exams = []
active_dates = {}
active_times = {}

r = 4
while r <= sheet.max_row:
    row_vals = [str(sheet.cell(r, c).value or '').strip() for c in range(1, sheet.max_column + 1)]
    if not any(row_vals):
        r += 1
        continue

    row_upper = [v.upper() for v in row_vals if v]
    if any('DATE' in v for v in row_upper) and any('TIME' in v for v in row_upper):
        for c in range(1, sheet.max_column + 1):
            val = str(sheet.cell(r, c).value or '').strip()
            if 'DATE' in val.upper():
                cur_date_col = c
            elif 'TIME' in val.upper():
                cur_time_col = c
            elif val and 'DATE' not in val.upper() and 'TIME' not in val.upper():
                venue_clean = re.sub(r'\s*\(\d+\)', '', val).strip()
                if c in venue_cols:
                    venue_cols[c]['venue'] = venue_clean
                    venue_cols[c]['date_col'] = cur_date_col
                    venue_cols[c]['time_col'] = cur_time_col
        r += 1
        continue

    # Update date & time
    for d_col in set(info['date_col'] for info in venue_cols.values() if info['date_col'] != -1):
        d_val = str(sheet.cell(r, d_col).value or '').strip()
        if d_val and 'DATE' not in d_val.upper():
            active_dates[d_col] = d_val

    for t_col in set(info['time_col'] for info in venue_cols.values() if info['time_col'] != -1):
        t_val = str(sheet.cell(r, t_col).value or '').strip()
        if t_val and 'TIME' not in t_val.upper():
            active_times[t_col] = t_val

    subject_row = r + 1

    for c, info in venue_cols.items():
        b_val = str(sheet.cell(r, c).value or '').strip()
        s_val = str(sheet.cell(subject_row, c).value or '').strip()

        if not b_val or not s_val:
            continue
        if b_val.upper() in ['BATCH', 'CLASS', 'DATE', 'TIME'] or s_val.upper() in ['SUBJECT', 'COURSE', 'DATE', 'TIME']:
            continue

        date = active_dates.get(info['date_col'], 'TBD')
        time = active_times.get(info['time_col'], '09:00 AM - 10:30 AM')

        all_exams.append({
            'row': r,
            'col': c,
            'date': date,
            'time': time,
            'room': info['venue'],
            'batch': b_val,
            'subject': s_val
        })

    r += 2

print(f"\n==========================================")
print(f"TOTAL MIDTERM EXAMS EXTRACTED: {len(all_exams)}")
print(f"==========================================")

# Check all batches
batches = sorted(list(set(e['batch'] for e in all_exams)))
print(f"\nTotal Unique Batches in Midterms: {len(batches)}")

# Check FA25-BCS (which in Spring 2026 is Semester 2 BCS!)
print("\n--- MIDTERMS FOR FA25-BCS (Semester 2 in Spring 2026) ---")
fa25_bcs = [e for e in all_exams if 'FA25-BCS' in e['batch'] or 'SP25-BCS' in e['batch']]
for e in fa25_bcs[:20]:
    print(f"  {e['date']} | {e['time']} | {e['room']} | {e['batch']} | {e['subject']}")

papers = set(f"{e['date']} - {e['subject']} ({e['batch']})" for e in fa25_bcs)
print(f"\nUnique Exam Papers for FA25/SP25 BCS in Midterms ({len(papers)}):")
for p in sorted(papers):
    print(f"  * {p}")
