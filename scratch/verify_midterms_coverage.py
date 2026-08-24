import openpyxl, re

file_path = 'assets/documents/Version I of Date Sheet Mid Term Exam SPRING 2026 25-03-2026.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb.active

venue_cols = {}
cur_date_col = -1
cur_time_col = -1

for c in range(1, sheet.max_column + 1):
    val = str(sheet.cell(3, c).value or '').strip()
    if 'DATE' in val.upper(): cur_date_col = c
    elif 'TIME' in val.upper() or 'SLOT' in val.upper(): cur_time_col = c
    elif val:
        venue_cols[c] = {'venue': re.sub(r'\s*\(\d+\)', '', val).strip(), 'date_col': cur_date_col, 'time_col': cur_time_col}

all_exams = []
active_dates = {}
active_times = {}

r = 4
while r <= sheet.max_row:
    row_vals = [str(sheet.cell(r, c).value or '').strip() for c in range(1, sheet.max_column + 1)]
    if not any(row_vals):
        r += 1
        continue
    row_upper = [v.upper() for v in row_vals if v]
    if any('DATE' in v for v in row_upper) and any('TIME' in v for v in row_upper):
        for c in range(1, sheet.max_column + 1):
            val = str(sheet.cell(r, c).value or '').strip()
            if 'DATE' in val.upper(): cur_date_col = c
            elif 'TIME' in val.upper() or 'SLOT' in val.upper(): cur_time_col = c
            elif val and 'DATE' not in val.upper() and 'TIME' not in val.upper():
                venue_clean = re.sub(r'\s*\(\d+\)', '', val).strip()
                if c in venue_cols:
                    venue_cols[c]['venue'] = venue_clean
                    venue_cols[c]['date_col'] = cur_date_col
                    venue_cols[c]['time_col'] = cur_time_col
        r += 1
        continue
    for d_col in set(info['date_col'] for info in venue_cols.values() if info['date_col'] != -1):
        d_val = str(sheet.cell(r, d_col).value or '').strip()
        if d_val and 'DATE' not in d_val.upper(): active_dates[d_col] = d_val
    for t_col in set(info['time_col'] for info in venue_cols.values() if info['time_col'] != -1):
        t_val = str(sheet.cell(r, t_col).value or '').strip()
        if t_val and 'TIME' not in t_val.upper(): active_times[t_col] = t_val
    subject_row = r + 1
    for c, info in venue_cols.items():
        b_val = str(sheet.cell(r, c).value or '').strip()
        s_val = str(sheet.cell(subject_row, c).value or '').strip()
        if not b_val or not s_val or b_val.upper() in ['BATCH', 'CLASS', 'DATE', 'TIME'] or s_val.upper() in ['SUBJECT', 'COURSE', 'DATE', 'TIME']:
            continue
        all_exams.append({
            'date': active_dates.get(info['date_col'], 'TBD'),
            'time': active_times.get(info['time_col'], '09:00 AM - 10:30 AM'),
            'room': info['venue'],
            'batch': b_val,
            'subject': s_val
        })
    r += 2

print(f"Total extracted Midterm records: {len(all_exams)}")

def is_match(user_batch, exam_batch):
    u = user_batch.upper()
    e = exam_batch.upper()
    if u in e: return True
    # If user is FA25-BCS-2-A, check FA25-BCS-A
    parts = u.split('-')
    if len(parts) == 4:
        short = f"{parts[0]}-{parts[1]}-{parts[3]}"
        if short in e: return True
    return False

test_batches = [
    'FA25-BCS-2-A', 'FA25-BCS-2-B', 'FA25-BCS-2-C',
    'FA24-BCS-4-A', 'FA23-BCS-6-A', 'FA22-BCS-8-A',
    'FA25-BSE-2-A', 'FA24-BSE-4-A', 'FA23-BSE-6-A',
    'FA24-BEE-4-A', 'FA25-BME-2-A', 'FA24-CVE-4-A'
]

for tb in test_batches:
    matched = [e for e in all_exams if is_match(tb, e['batch'])]
    unique_papers = sorted(list(set(f"{e['date']} - {e['subject']}" for e in matched)))
    print(f"\n{tb} -> {len(matched)} entries, {len(unique_papers)} unique papers:")
    for p in unique_papers:
        print(f"    * {p}")
