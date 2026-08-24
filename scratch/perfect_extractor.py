import openpyxl
import json
import re

wb = openpyxl.load_workbook('assets/documents/Date_Sheet_FINAL_Term_Exam_FALL_2025 - Version I - 08-12-25.xlsx', data_only=True)
sheet = wb.active

# 1. Identify venue columns in row 3
venue_cols = {}
cur_date_col = -1
cur_time_col = -1

for c in range(1, sheet.max_column + 1):
    val = str(sheet.cell(3, c).value or '').strip()
    val_upper = val.upper()
    if 'DATE' in val_upper:
        cur_date_col = c
    elif 'TIME' in val_upper:
        cur_time_col = c
    elif val:
        venue_clean = re.sub(r'\s*\(\d+\)', '', val).strip()
        venue_cols[c] = {
            'venue': venue_clean,
            'date_col': cur_date_col,
            'time_col': cur_time_col
        }

print(f"Venue columns mapped: {len(venue_cols)}")

all_exams = []
active_dates = {}
active_times = {}

r = 4
while r <= sheet.max_row:
    # Check if row is empty
    row_vals = [str(sheet.cell(r, c).value or '').strip() for c in range(1, sheet.max_column + 1)]
    if not any(row_vals):
        r += 1
        continue

    # Check if row is a repeated header
    row_upper = [v.upper() for v in row_vals if v]
    if any('DATE' in v for v in row_upper) and any('TIME' in v for v in row_upper):
        # Update venue headers if needed
        for c in range(1, sheet.max_column + 1):
            val = str(sheet.cell(r, c).value or '').strip()
            if 'DATE' in val.upper():
                cur_date_col = c
            elif 'TIME' in val.upper():
                cur_time_col = c
            elif val and 'DATE' not in val.upper() and 'TIME' not in val.upper():
                venue_clean = re.sub(r'\s*\(\d+\)', '', val).strip()
                if c in venue_cols:
                    venue_cols[c]['venue'] = venue_clean
                    venue_cols[c]['date_col'] = cur_date_col
                    venue_cols[c]['time_col'] = cur_time_col
        r += 1
        continue

    # Update date & time from this row
    for d_col in set(info['date_col'] for info in venue_cols.values() if info['date_col'] != -1):
        d_val = str(sheet.cell(r, d_col).value or '').strip()
        if d_val and 'DATE' not in d_val.upper():
            active_dates[d_col] = d_val

    for t_col in set(info['time_col'] for info in venue_cols.values() if info['time_col'] != -1):
        t_val = str(sheet.cell(r, t_col).value or '').strip()
        if t_val and 'TIME' not in t_val.upper():
            active_times[t_col] = t_val

    # Find the subject row for this batch row (it's the next row r+1)
    subject_row = r + 1
    
    # Extract exams
    for c, info in venue_cols.items():
        b_val = str(sheet.cell(r, c).value or '').strip()
        s_val = str(sheet.cell(subject_row, c).value or '').strip()

        if not b_val or not s_val:
            continue
        if b_val.upper() in ['BATCH', 'CLASS', 'DATE', 'TIME'] or s_val.upper() in ['SUBJECT', 'COURSE', 'DATE', 'TIME']:
            continue

        date = active_dates.get(info['date_col'], 'TBD')
        time = active_times.get(info['time_col'], '09:00 AM - 12:00 PM')

        all_exams.append({
            'row': r,
            'col': c,
            'date': date,
            'time': time,
            'room': info['venue'],
            'batch': b_val,
            'subject': s_val
        })

    # Advance past the subject row
    r += 2

print(f"\n==========================================")
print(f"TOTAL EXTRACTED EXAMS: {len(all_exams)}")
print(f"==========================================")

# Now check for FA25-BCS-A
print("\n--- ALL EXAMS FOR FA25-BCS-A ---")
fa25_a = [e for e in all_exams if 'FA25-BCS-A' in e['batch']]
for e in fa25_a:
    print(f"  {e['date']} | {e['time']} | {e['room']} | {e['batch']} | {e['subject']}")

# Unique papers
papers = {}
for e in fa25_a:
    key = f"{e['date']} | {e['time']} | {e['subject']}"
    if key not in papers:
        papers[key] = []
    papers[key].append(e['room'])

print(f"\nUnique Exam Papers for FA25-BCS-A ({len(papers)}):")
for p, rooms in papers.items():
    print(f"  * {p} (Rooms: {', '.join(rooms)})")
