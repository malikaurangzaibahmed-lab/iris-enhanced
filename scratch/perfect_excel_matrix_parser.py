import openpyxl
import json
import re

def parse_quad_excel(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active
    
    # 1. Identify venue header columns and date/time columns in row 3
    # Look for repeating header blocks across columns
    venue_cols = {} # col_idx -> {'venue': str, 'date_col': int, 'time_col': int}
    
    cur_date_col = -1
    cur_time_col = -1
    
    for c in range(1, sheet.max_column + 1):
        val = str(sheet.cell(3, c).value or '').strip()
        val_upper = val.upper()
        if 'DATE' in val_upper:
            cur_date_col = c
        elif 'TIME' in val_upper:
            cur_time_col = c
        elif val:
            venue_clean = re.sub(r'\s*\(\d+\)', '', val).strip() # "C 1.1 (49)" -> "C 1.1"
            venue_cols[c] = {
                'venue': venue_clean,
                'raw_venue': val,
                'date_col': cur_date_col,
                'time_col': cur_time_col
            }
            
    print(f"Mapped {len(venue_cols)} venue columns across blocks.")
    
    all_exams = []
    
    # Track current active date per date_col
    active_dates = {}
    active_times = {}
    
    r = 4
    while r <= sheet.max_row:
        # Check if this row is a repeated header row (e.g. Row 42 has "Date", "Time", etc.)
        row_first_vals = [str(sheet.cell(r, c).value or '').strip().upper() for c in range(1, min(sheet.max_column, 15))]
        if any('DATE' in v for v in row_first_vals) and any('TIME' in v for v in row_first_vals):
            # Header row repeated in middle of sheet: skip to next data row
            r += 1
            continue
            
        # Update active dates and times from this row
        for d_col in set(info['date_col'] for info in venue_cols.values() if info['date_col'] != -1):
            d_val = str(sheet.cell(r, d_col).value or '').strip()
            # If row has a date e.g. "Tuesday 30-12-2025", update active date
            if d_val and 'DATE' not in d_val.upper():
                active_dates[d_col] = d_val
                
        for t_col in set(info['time_col'] for info in venue_cols.values() if info['time_col'] != -1):
            t_val = str(sheet.cell(r, t_col).value or '').strip()
            if t_val and 'TIME' not in t_val.upper():
                active_times[t_col] = t_val
                
        # Now extract exam pairs for each venue column
        for c, info in venue_cols.items():
            batch_raw = str(sheet.cell(r, c).value or '').strip()
            subject_raw = str(sheet.cell(r + 1, c).value or '').strip()
            
            if not batch_raw or not subject_raw:
                continue
            if batch_raw.upper() in ['BATCH', 'CLASS', 'DATE', 'TIME'] or subject_raw.upper() in ['SUBJECT', 'COURSE', 'DATE', 'TIME']:
                continue
                
            date = active_dates.get(info['date_col'], 'TBD')
            time = active_times.get(info['time_col'], '09:00 AM - 12:00 PM')
            
            # Format time e.g. "0100-0400" -> "01:00 PM - 04:00 PM", "0900-1200" -> "09:00 AM - 12:00 PM"
            time_formatted = format_slot_time(time)
            
            # Split compound subjects/batches if needed or expand sections
            all_exams.append({
                'date': date,
                'time': time_formatted,
                'room': info['venue'],
                'batch': batch_raw,
                'subject': subject_raw
            })
            
        r += 2
        
    return all_exams

def format_slot_time(raw):
    raw = raw.strip()
    if '-' in raw:
        parts = raw.split('-')
        if len(parts) == 2 and len(parts[0].strip()) == 4 and len(parts[1].strip()) == 4:
            s_h = int(parts[0][:2])
            s_m = parts[0][2:]
            e_h = int(parts[1][:2])
            e_m = parts[1][2:]
            
            s_ampm = 'PM' if s_h >= 12 or s_h in [1,2,3,4,5] else 'AM'
            if s_h == 1: s_h = 13 # 0100 is 1 PM
            if s_h == 2: s_h = 14
            if s_h == 3: s_h = 15
            if s_h == 4: s_h = 16
            
            s_disp = s_h - 12 if s_h > 12 else (s_h if s_h > 0 else 12)
            e_disp = e_h - 12 if e_h > 12 else (e_h if e_h > 0 else 12)
            e_ampm = 'PM' if e_h >= 12 or e_h in [1,2,3,4,5] else 'AM'
            
            return f"{s_disp:02d}:{s_m} {s_ampm} - {e_disp:02d}:{e_m} {e_ampm}"
    return raw

exams = parse_quad_excel('assets/documents/Date_Sheet_FINAL_Term_Exam_FALL_2025 - Version I - 08-12-25.xlsx')
print(f"\nExtracted TOTAL {len(exams)} exam records.")

# Check for FA25-BCS-A
print("\n=== FA25-BCS-A EXAMS ===")
fa25_a = [e for e in exams if 'FA25-BCS-A' in e['batch']]
for e in fa25_a:
    print(f"  {e['date']} | {e['time']} | {e['room']} | {e['batch']} | {e['subject']}")

# Group by date & subject to verify unique papers
papers = {}
for e in fa25_a:
    key = f"{e['date']} - {e['subject']}"
    if key not in papers:
        papers[key] = []
    papers[key].append(e['room'])

print(f"\nUnique Exam Papers for FA25-BCS-A ({len(papers)}):")
for p, rooms in papers.items():
    print(f"  ✓ {p} (Rooms: {', '.join(rooms)})")
