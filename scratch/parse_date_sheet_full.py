import openpyxl
import json
import re

wb = openpyxl.load_workbook('assets/documents/Date_Sheet_FINAL_Term_Exam_FALL_2025 - Version I - 08-12-25.xlsx', data_only=True)
sheet = wb.active

# 1. Identify blocks and date/time columns in row 3
col_info = {} # col_idx -> {'venue': str, 'date_col': int, 'time_col': int}

cur_date_col = -1
cur_time_col = -1

for c in range(1, sheet.max_column + 1):
    val = str(sheet.cell(3, c).value or '').strip()
    val_upper = val.upper()
    if 'DATE' in val_upper:
        cur_date_col = c
    elif 'TIME' in val_upper:
        cur_time_col = c
    elif val:
        col_info[c] = {
            'venue': val,
            'date_col': cur_date_col,
            'time_col': cur_time_col
        }

print(f"Total venue columns mapped: {len(col_info)}")

all_exams = []

# Scan pairs of rows (r = batch row, r+1 = subject row)
# Dates and times are set on batch row or subject row, or carried down
last_dates = {}
last_times = {}

r = 4
while r <= sheet.max_row:
    for c, info in col_info.items():
        d_col = info['date_col']
        t_col = info['time_col']
        
        # Check date in row r or r+1 or carried
        d_val = str(sheet.cell(r, d_col).value or sheet.cell(r+1, d_col).value or '').strip()
        if d_val:
            last_dates[d_col] = d_val
        date = last_dates.get(d_col, '')
        
        t_val = str(sheet.cell(r, t_col).value or sheet.cell(r+1, t_col).value or '').strip()
        if t_val:
            last_times[t_col] = t_val
        time = last_times.get(t_col, '')
        
        batch = str(sheet.cell(r, c).value or '').strip()
        subject = str(sheet.cell(r+1, c).value or '').strip()
        
        if batch and subject and batch.upper() != 'BATCH' and subject.upper() != 'SUBJECT':
            all_exams.append({
                'row': r,
                'col': c,
                'date': date,
                'time': time,
                'venue': info['venue'],
                'batch': batch,
                'subject': subject
            })
    r += 2

print(f"Total extracted exam entries: {len(all_exams)}")

# Let's inspect all batches present in the Excel
batches = sorted(list(set(e['batch'] for e in all_exams)))
print(f"\nUnique batch strings in Date Sheet ({len(batches)}):")
for b in batches:
    print(f"  {b}")

# Let's search for FA25-BCS, BCS-2, FA24-BCS, etc.
print("\n--- Exams for FA25-BCS batches ---")
fa25_bcs = [e for e in all_exams if 'FA25-BCS' in e['batch'].upper() or 'BCS-2' in e['batch'].upper() or 'BCS-A' in e['batch'].upper()]
for e in fa25_bcs:
    print(f"{e['date']} | {e['time']} | {e['venue']} | {e['batch']} | {e['subject']}")
