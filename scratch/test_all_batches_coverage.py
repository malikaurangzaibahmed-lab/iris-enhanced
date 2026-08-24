import re

# Simulate the Dart BatchKey.isBatchMatch & FormatGuard.expandBatchSections logic
def expand_batch_sections(raw):
    if not raw or not raw.strip():
        return []
    clean = raw.strip().upper()
    chunks = [c.strip() for c in re.split(r'[,&/]', clean) if c.strip()]
    if len(chunks) > 1:
        results = []
        for chunk in chunks:
            if '-' in chunk:
                results.append(chunk)
            elif results and '-' in results[-1]:
                prev = results[-1]
                parts = prev.split('-')
                if len(parts) >= 2:
                    parts[-1] = chunk
                    results.append('-'.join(parts))
                else:
                    results.append(chunk)
            else:
                results.append(chunk)
        return results

    if '-' in clean:
        parts = clean.split('-')
        last_part = parts[-1]
        range_match = re.match(r'^([A-Za-z0-9]+)\s*-\s*([A-Za-z0-9]+)$', last_part)
        if range_match:
            base = '-'.join(parts[:-1])
            return [f"{base}-{range_match.group(1)}", f"{base}-{range_match.group(2)}"]
        if re.search(r'[A-Za-z]\s*-\s*[A-Za-z]', clean):
            c_parts = clean.split('-')
            if len(c_parts) == 4:
                base = f"{c_parts[0]}-{c_parts[1]}"
                return [f"{base}-{c_parts[2]}", f"{base}-{c_parts[3]}"]

    return [clean]

def normalize_prog(p):
    p = p.upper()
    if p in ['BCS', 'BSCS', 'CS']: return 'BCS'
    if p in ['BSE', 'BSSE', 'SE']: return 'BSE'
    if p in ['BEE', 'BSEE', 'EE']: return 'BEE'
    if p in ['BBA', 'BBS']: return 'BBA'
    if p in ['BME', 'BSME', 'ME']: return 'BME'
    if p in ['CVE', 'BCE', 'CE']: return 'CVE'
    return p

def parse_batch_key(raw):
    clean = raw.strip().upper()
    parts = [p.strip() for p in clean.split('-') if p.strip()]
    if not parts:
        return {'intake': '', 'program': clean, 'semester': 1, 'section': 'A'}
    if len(parts) == 1:
        return {'intake': '', 'program': normalize_prog(parts[0]), 'semester': 1, 'section': 'A'}
    if len(parts) == 2:
        if re.match(r'^(?:FA|SP)\d{2}$', parts[0]):
            return {'intake': parts[0], 'program': normalize_prog(parts[1]), 'semester': 1, 'section': 'A'}
        m = re.match(r'^(\d+)([A-Z]*)', parts[1])
        if m:
            return {'intake': '', 'program': normalize_prog(parts[0]), 'semester': int(m.group(1)), 'section': m.group(2) or 'A'}
        return {'intake': '', 'program': normalize_prog(parts[0]), 'semester': 1, 'section': parts[1]}
    
    intake = parts[0]
    prog = normalize_prog(parts[1])
    if len(parts) >= 4:
        sem = int(parts[2]) if parts[2].isdigit() else 1
        sec = parts[3]
        return {'intake': intake, 'program': prog, 'semester': sem, 'section': sec}
    
    last = parts[2]
    m = re.match(r'^(\d+)([A-Z]*)', last)
    if m:
        return {'intake': intake, 'program': prog, 'semester': int(m.group(1)), 'section': m.group(2) or 'A'}
    return {'intake': intake, 'program': prog, 'semester': 1, 'section': last}

def is_batch_match(student_raw, exam_raw):
    if not student_raw or not exam_raw: return False
    s_clean = student_raw.strip().upper()
    e_clean = exam_raw.strip().upper()
    if s_clean == e_clean or s_clean == 'ALL' or e_clean == 'ALL': return True
    
    s_expanded = expand_batch_sections(student_raw)
    e_expanded = expand_batch_sections(exam_raw)
    
    for eb in e_expanded:
        for sb in s_expanded:
            if eb.strip().upper() == sb.strip().upper():
                return True
            s_key = parse_batch_key(sb)
            e_key = parse_batch_key(eb)
            
            if s_key['program'] != e_key['program']:
                continue
            sec_match = not s_key['section'] or not e_key['section'] or s_key['section'] == e_key['section']
            
            intake_match = True
            if s_key['intake'] and e_key['intake']:
                intake_match = s_key['intake'] == e_key['intake']
                
            if sec_match and intake_match:
                return True
    return False

# Now let's test all extracted exams from our perfect parser
import subprocess
out = subprocess.check_output(['python', 'scratch/perfect_extractor.py'], text=True)
print("Extractor output ready.")

# Check matches for multiple batches
test_batches = [
    'FA25-BCS-2-A', 'FA25-BCS-A', 'BCS-2A',
    'FA25-BCS-2-B', 'FA25-BCS-B',
    'FA25-BCS-2-C', 'FA25-BCS-C',
    'FA24-BCS-2-A', 'FA24-BCS-A',
    'FA24-BSE-2-A', 'FA24-BSE-A',
    'FA23-BCS-4-A', 'FA23-BCS-A',
    'FA22-BCS-6-A', 'FA22-BCS-A'
]

# Let's import the exams list from perfect_extractor
import openpyxl
wb = openpyxl.load_workbook('assets/documents/Date_Sheet_FINAL_Term_Exam_FALL_2025 - Version I - 08-12-25.xlsx', data_only=True)
sheet = wb.active

venue_cols = {}
cur_date_col = -1
cur_time_col = -1
for c in range(1, sheet.max_column + 1):
    val = str(sheet.cell(3, c).value or '').strip()
    if 'DATE' in val.upper(): cur_date_col = c
    elif 'TIME' in val.upper(): cur_time_col = c
    elif val:
        venue_cols[c] = {'venue': re.sub(r'\s*\(\d+\)', '', val).strip(), 'date_col': cur_date_col, 'time_col': cur_time_col}

all_exams = []
active_dates = {}
active_times = {}
r = 4
while r <= sheet.max_row:
    row_vals = [str(sheet.cell(r, c).value or '').strip() for c in range(1, sheet.max_column + 1)]
    if not any(row_vals):
        r += 1
        continue
    row_upper = [v.upper() for v in row_vals if v]
    if any('DATE' in v for v in row_upper) and any('TIME' in v for v in row_upper):
        r += 1
        continue
    for d_col in set(info['date_col'] for info in venue_cols.values() if info['date_col'] != -1):
        d_val = str(sheet.cell(r, d_col).value or '').strip()
        if d_val and 'DATE' not in d_val.upper(): active_dates[d_col] = d_val
    for t_col in set(info['time_col'] for info in venue_cols.values() if info['time_col'] != -1):
        t_val = str(sheet.cell(r, t_col).value or '').strip()
        if t_val and 'TIME' not in t_val.upper(): active_times[t_col] = t_val
    subject_row = r + 1
    for c, info in venue_cols.items():
        b_val = str(sheet.cell(r, c).value or '').strip()
        s_val = str(sheet.cell(subject_row, c).value or '').strip()
        if not b_val or not s_val or b_val.upper() in ['BATCH', 'CLASS', 'DATE', 'TIME'] or s_val.upper() in ['SUBJECT', 'COURSE', 'DATE', 'TIME']:
            continue
        all_exams.append({
            'date': active_dates.get(info['date_col'], 'TBD'),
            'time': active_times.get(info['time_col'], '09:00 AM - 12:00 PM'),
            'room': info['venue'],
            'batch': b_val,
            'subject': s_val
        })
    r += 2

print(f"\nTotal exams in dataset: {len(all_exams)}")

for tb in test_batches:
    matched = [e for e in all_exams if is_batch_match(tb, e['batch'])]
    unique_subjects = set(f"{e['date']} - {e['subject']}" for e in matched)
    print(f"\nBatch: {tb:15} -> {len(matched)} exam session rows, {len(unique_subjects)} unique exam papers:")
    for us in sorted(unique_subjects):
        print(f"    * {us}")
