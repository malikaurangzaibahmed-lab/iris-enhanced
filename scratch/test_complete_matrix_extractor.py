import openpyxl
import json
import re

wb = openpyxl.load_workbook('assets/documents/Date_Sheet_FINAL_Term_Exam_FALL_2025 - Version I - 08-12-25.xlsx', data_only=True)
sheet = wb.active

# Map columns
venue_cols = {}
cur_date_col = -1
cur_time_col = -1

for c in range(1, sheet.max_column + 1):
    val = str(sheet.cell(3, c).value or '').strip()
    val_upper = val.upper()
    if 'DATE' in val_upper:
        cur_date_col = c
    elif 'TIME' in val_upper:
        cur_time_col = c
    elif val:
        venue_clean = re.sub(r'\s*\(\d+\)', '', val).strip()
        venue_cols[c] = {
            'venue': venue_clean,
            'date_col': cur_date_col,
            'time_col': cur_time_col
        }

all_exams = []
active_dates = {}
active_times = {}

r = 4
while r <= sheet.max_row:
    # Check if row is a repeated header
    row_headers = [str(sheet.cell(r, c).value or '').strip().upper() for c in range(1, min(sheet.max_column, 20))]
    if any('DATE' in h for h in row_headers) and any('TIME' in h for h in row_headers):
        # Update venue headers if any changed
        for c in range(1, sheet.max_column + 1):
            val = str(sheet.cell(r, c).value or '').strip()
            if 'DATE' in val.upper():
                cur_date_col = c
            elif 'TIME' in val.upper():
                cur_time_col = c
            elif val and 'DATE' not in val.upper() and 'TIME' not in val.upper():
                venue_clean = re.sub(r'\s*\(\d+\)', '', val).strip()
                if c in venue_cols:
                    venue_cols[c]['venue'] = venue_clean
                    venue_cols[c]['date_col'] = cur_date_col
                    venue_cols[c]['time_col'] = cur_time_col
        r += 1
        continue

    # Update date & time
    for d_col in set(info['date_col'] for info in venue_cols.values() if info['date_col'] != -1):
        d_val = str(sheet.cell(r, d_col).value or '').strip()
        if d_val and 'DATE' not in d_val.upper():
            active_dates[d_col] = d_val

    for t_col in set(info['time_col'] for info in venue_cols.values() if info['time_col'] != -1):
        t_val = str(sheet.cell(r, t_col).value or '').strip()
        if t_val and 'TIME' not in t_val.upper():
            active_times[t_col] = t_val

    # Read batch & subject
    for c, info in venue_cols.items():
        b_val = str(sheet.cell(r, c).value or '').strip()
        s_val = str(sheet.cell(r + 1, c).value or '').strip()

        if not b_val or not s_val:
            continue
        if b_val.upper() in ['BATCH', 'CLASS', 'DATE', 'TIME'] or s_val.upper() in ['SUBJECT', 'COURSE', 'DATE', 'TIME']:
            continue

        date = active_dates.get(info['date_col'], 'TBD')
        time = active_times.get(info['time_col'], '0900-1200')

        all_exams.append({
            'row': r,
            'col': c,
            'date': date,
            'time': time,
            'room': info['venue'],
            'batch': b_val,
            'subject': s_val
        })

    r += 2

print(f"Total extracted: {len(all_exams)}")

# Let's inspect FA25-BCS-A
print("\nALL FA25-BCS-A ENTRIES:")
for e in all_exams:
    if 'FA25-BCS-A' in e['batch']:
        print(f"Row {e['row']}, Col {e['col']} -> Date: {e['date']} | Time: {e['time']} | Room: {e['room']} | Batch: {e['batch']} | Subject: {e['subject']}")
